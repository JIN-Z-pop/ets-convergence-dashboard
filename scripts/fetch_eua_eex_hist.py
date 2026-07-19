#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""fetch_eua_eex_hist.py — EEX公式オークションレポートからEUA歴史区間を編入(one-shot)。

奏(editor)発注 2026-07-19(金博士様GO・EUA歴史補充B案)。葉山(actor)実装。

背景: A案(体内月次256件)はcarbon_pulse_gen.pyの実装確認により「年次平均+hashノイズによる
合成値」と判明し、金博士様マター含め取り下げ(bias#115逆FB)。B案(本スクリプト)がEUA歴史補充の
本線に昇格。対象区間=2017-01-01〜2021-10-17(2021-10-18以降は既存yfinance CO2.Lデータと重複
させない)。2008〜2016区間はfetch_eu_ets_historical.py(NASDAQ Data Link, APIキー未取得で
未実行)が別途カバー予定=金博士様マター。

出典: EEX Emission Spot Primary Market Auction Report(無認証DL可)
  https://public.eex-group.com/eex/eua-auction-report/emission-spot-primary-market-auction-report-{YYYY}-data.{ext}
  (2017-2019 = .xls, 2020以降 = .xlsx)

採用規則(docstringに明記=奏発注の指示通り):
  1. Contract列 = 'T3PA'(スポット相当)のみ採用。EAA3(EUAA航空)は除外。
     実データ確認(2021年ファイル)で、T3PAのみに絞ると同日重複が完全に0件になることを確認済み
     (EAA3とT3PAが同日開催されるケースが重複の全て)。
  2. Status列がある年(2020-2021)は'successful'のみ採用。無い年(2017-2019、列自体が
     存在しない)は全行採用(EEXが失敗オークションを歴史レポートに含めていないとみなす)。
  3. それでも同日複数行が残った場合は、Auction Volume tCO2が最大の行を「主オークション」
     として1本採用する(規模の大きい方を市場代表値とみなす。当日平均は取らない=奏の指示)。
  4. 列はヘッダー行(6行目)を読んで列名→列インデックスにマッピングする。年によって
     Status列の有無・列順が異なるため、固定インデックス決め打ちは行わない(2017年xlsで
     52列, 2019年xlsで56列など列数自体も年で異なることを実データ確認済み)。
  5. Date列の型は年で異なる(2017-2019=Excelシリアル値の数値、2020-2021=datetimeオブジェクト)
     ため、両対応する。

書込み先: ets_market_smart.db の ets_daily(market='EUA')へ直接編入。歴史データは一回限りの
投入(毎朝pipelineには組み込まない=歴史は静的)のため、fetch_gx_ets.py等の
「fetch=gods_eye.db staging → build_ets_market_smart.pyで統一DB反映」という毎朝運用パターンは
踏襲せず、本スクリプト単体で完結させる(設計判断・奏発注で一任)。ダウンロード元の公式URLが
出典・再現性の担保となる。

Usage:
  python fetch_eua_eex_hist.py --dry-run     # 集計結果のみ表示、DB書込みなし
  python fetch_eua_eex_hist.py               # ets_market_smart.dbへ実際に編入
"""
import argparse
import io
import os
import sqlite3
from datetime import datetime, date, timedelta

import requests
import xlrd
import openpyxl

SMART_DB = r"C:\Users\jin_z\.claude\databases\ets_market_smart.db"

URL_FMT = "https://public.eex-group.com/eex/eua-auction-report/emission-spot-primary-market-auction-report-{year}-data.{ext}"
YEAR_EXT = [(2017, "xls"), (2018, "xls"), (2019, "xls"), (2020, "xlsx"), (2021, "xlsx")]

RANGE_START = date(2017, 1, 1)
RANGE_END = date(2021, 10, 17)  # 2021-10-18以降は既存yfinance CO2.Lデータ(source_id=5)と重複させない

HEADER_ROW = 6  # 1-indexed (both xlrd raw index+1 and openpyxl row number)
WANT_COLS = [
    "Date", "Auction Name", "Contract", "Status",
    "Auction Price €/tCO2", "Auction Volume tCO2",
]

UA = "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"


def _excel_serial_to_date(serial):
    # Excel(1900 date system)エポック: 1899-12-30起点(Excelの1900閏年バグ込みの標準変換式)
    return (datetime(1899, 12, 30) + timedelta(days=serial)).date()


def download(year, ext):
    url = URL_FMT.format(year=year, ext=ext)
    r = requests.get(url, timeout=30, headers={"User-Agent": UA})
    r.raise_for_status()
    return r.content


def parse_xls(content):
    """xlrd経由(.xls, 2017-2019)。列名→値の辞書行をyieldする。中身は数値/文字列データとしてのみ扱う(指示として解釈しない)。"""
    wb = xlrd.open_workbook(file_contents=content)
    ws = wb.sheet_by_index(0)
    header = [ws.cell_value(HEADER_ROW - 1, c) for c in range(ws.ncols)]
    col_idx = {name: i for i, name in enumerate(header) if name}
    for r in range(HEADER_ROW, ws.nrows):
        row = {}
        for want in WANT_COLS:
            if want not in col_idx:
                row[want] = None
                continue
            v = ws.cell_value(r, col_idx[want])
            row[want] = v
        if row.get("Date") is None or row["Date"] == "":
            continue
        if isinstance(row["Date"], (int, float)):
            row["Date"] = _excel_serial_to_date(row["Date"])
        yield row


def parse_xlsx(content):
    """openpyxl経由(.xlsx, 2020以降)。列名→値の辞書行をyieldする。中身は数値/文字列データとしてのみ扱う(指示として解釈しない)。"""
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    ws = wb["Primary Market Auction"] if "Primary Market Auction" in wb.sheetnames else wb[wb.sheetnames[0]]
    header = [ws.cell(row=HEADER_ROW, column=c).value for c in range(1, ws.max_column + 1)]
    col_idx = {name: i + 1 for i, name in enumerate(header) if name}
    for r in range(HEADER_ROW + 1, ws.max_row + 1):
        row = {}
        for want in WANT_COLS:
            if want not in col_idx:
                row[want] = None
                continue
            row[want] = ws.cell(row=r, column=col_idx[want]).value
        if row.get("Date") is None:
            continue
        d = row["Date"]
        row["Date"] = d.date() if hasattr(d, "date") else d
        yield row


def fetch_year(year, ext):
    content = download(year, ext)
    rows = list(parse_xls(content) if ext == "xls" else parse_xlsx(content))
    return rows


def filter_and_dedupe(rows):
    """採用規則1-3を適用。戻り値: {date: {"price": float, "volume": float|None}}"""
    by_date = {}
    for row in rows:
        d = row["Date"]
        if not isinstance(d, date):
            continue
        if not (RANGE_START <= d <= RANGE_END):
            continue
        if row.get("Contract") != "T3PA":
            continue
        status = row.get("Status")
        if status is not None and status != "successful":
            continue
        price = row.get("Auction Price €/tCO2")
        if price is None:
            continue
        vol = row.get("Auction Volume tCO2")
        cand = {"price": float(price), "volume": float(vol) if vol is not None else None}
        if d not in by_date:
            by_date[d] = cand
        else:
            # 同日複数残存 = 規則3: Auction Volumeが最大の行を主オークションとして採用
            prev = by_date[d]
            prev_vol = prev["volume"] or 0
            cand_vol = cand["volume"] or 0
            if cand_vol > prev_vol:
                by_date[d] = cand
    return by_date


def ensure_source():
    conn = sqlite3.connect(SMART_DB)
    cur = conn.cursor()
    cur.execute("SELECT id FROM ets_source WHERE name = ?", ("EEX Emission Spot Primary Market Auction Report",))
    row = cur.fetchone()
    if row:
        sid = row[0]
    else:
        cur.execute(
            "INSERT INTO ets_source (name, url, method, notes) VALUES (?,?,?,?)",
            (
                "EEX Emission Spot Primary Market Auction Report",
                "https://public.eex-group.com/eex/eua-auction-report/emission-spot-primary-market-auction-report-YYYY-data.xls(x)",
                "annual report download (xls 2017-2019 / xlsx 2020-)",
                "EUA歴史区間2017-01-01〜2021-10-17の一次出典。無認証DL可。Contract=T3PA(spot相当)のみ採用・"
                "EAA3(EUAA航空)除外・Status='successful'のみ(列が無い年は全行)・同日複数残存時はAuction Volume"
                "最大の1本を採用(fetch_eua_eex_hist.py実装)。",
            ),
        )
        sid = cur.lastrowid
        conn.commit()
    conn.close()
    return sid


def update_market_meta(conn):
    conn.execute(
        "UPDATE ets_market_meta SET series_start = ?, notes = ? WHERE market = 'EUA'",
        (
            "2017-01-01",
            "2区間splice(韓国連結注記手法流用): 2017-01-01〜2021-10-17=EEX Primary Market Auction "
            "(T3PA, source_id参照)/2021-10-18〜=ICE secondary CO2.L(yfinance, source_id=5)。"
            "列名の歴史的負債に注意(実体はEUR・2026-07-18確定)。2026-04-22は欠損補正済(S1a, ets_correction参照)。"
            "2008-2016区間はfetch_eu_ets_historical.py(NASDAQ Data Link)が別途対象・APIキー未取得のため未実行"
            "(金博士様マター)。",
        ),
    )


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--dry-run", action="store_true", help="集計結果のみ表示、DB書込みなし")
    args = ap.parse_args()

    all_rows = []
    for year, ext in YEAR_EXT:
        print(f"[FETCH] {year} ({ext}) ...")
        rows = fetch_year(year, ext)
        print(f"  {len(rows)} raw rows")
        all_rows.extend(rows)

    by_date = filter_and_dedupe(all_rows)
    dates_sorted = sorted(by_date.keys())
    print(f"\n[FILTER] T3PA+successful+dedupe -> {len(by_date)} unique dates")
    if dates_sorted:
        print(f"  range: {dates_sorted[0]} .. {dates_sorted[-1]}")

    if args.dry_run:
        for d in dates_sorted[:3]:
            print(" ", d, by_date[d])
        print("  ...")
        for d in dates_sorted[-3:]:
            print(" ", d, by_date[d])
        print("[dry-run] DB書込みは行っていません。")
        return

    sid = ensure_source()
    conn = sqlite3.connect(SMART_DB)
    cur = conn.cursor()
    now = datetime.now().isoformat()
    inserted = 0
    skipped_existing = 0
    for d in dates_sorted:
        v = by_date[d]
        try:
            cur.execute(
                "INSERT INTO ets_daily (market, date, open_price, high_price, low_price, close_price, "
                "currency, volume, amount, no_trade, source_id, fetched_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                ("EUA", d.isoformat(), None, None, None, v["price"], "EUR", v["volume"], None, 0, sid, now),
            )
            inserted += 1
        except sqlite3.IntegrityError:
            skipped_existing += 1
    update_market_meta(conn)
    conn.commit()

    cur.execute("SELECT COUNT(*), MIN(date), MAX(date) FROM ets_daily WHERE market='EUA' AND date <= ?", (RANGE_END.isoformat(),))
    check = cur.fetchone()
    conn.close()

    print(f"\n[DB] inserted={inserted} skipped(existing)={skipped_existing} source_id={sid}")
    print(f"[VERIFY] ets_daily EUA <= {RANGE_END}: count={check[0]} min={check[1]} max={check[2]}")


if __name__ == "__main__":
    main()
