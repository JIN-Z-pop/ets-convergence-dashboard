#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""load_eua_eex_auction_hist.py — EUA一次市場オークション結果(EEX)を
gods_eye.dbの恒久層テーブル raw_eua_auction_eex へ投入する、一回限りの投入スクリプト。

【恒久層投入・build側が毎朝参照】このテーブルは gods_eye.db 側の恒久データ。
ets_market_smart.db.ets_auction は build_ets_market_smart.py が毎朝DELETE→
全再構築する揮発層 — 直接そちら側へINSERTしてはならない(2026-07-19の教訓、
EEXデータを直接ets_dailyへ投入し翌rebuildで消滅した事故を繰り返さないこと)。

奏(editor) spec: eua_eex_auction_load_spec_20260720.md / 葉山(actor)実装。

投入元: data/sources/eua_hist/eex_auction_{2017..2026}.{xls|xlsx} 全10ファイル
        (md5はSOURCES.mdに記録済み・投入前に照合)
ヘッダ行: Excel第6行だが固定index禁止 — 「Date」「Auction Name」を両方含む行を動的検出。
列: 位置ベース禁止・列名ベースでmapping(列数は52/57/61と年により変動)。
スコープ: spec §3スキーマに列挙された列のみ投入。bids詳細・per-bidder統計・国別配分列群は
          spec §5 V5(a)により意図的に対象外(素材ファイル自体は恒久保全済み・将来再parse可)。

Usage: python scripts/load_eua_eex_auction_hist.py  (ets-convergence-dashboardルートから)
"""
import hashlib
import sqlite3
from datetime import datetime
from pathlib import Path

import xlrd
import openpyxl

GODS = r"C:\Users\jin_z\.claude\databases\gods_eye.db"
SRC_DIR = Path(r"C:\Users\jin_z\Desktop\ets-convergence-dashboard\data\sources\eua_hist")

EXPECTED_MD5 = {
    2017: "f2c6aec006feb3064361f774a873c54e",
    2018: "5b762ec36dd5762ad944ce9e14ce9a8e",
    2019: "13a32ec274dc4ea95088d17c8a70db45",
    2020: "7ab0fba25132e4de4bbd24d599310575",
    2021: "212bd468d5e96ba2b70ba33c3a30bef8",
    2022: "96fe7f308993999fc44b23a51e8a9fd8",
    2023: "ff48988ac3090f965eaa88b12369d844",
    2024: "a13fe5a985e8f997fca4a0e1d1798435",
    2025: "2db5aec3ee0708e0819b689a68aab47e",
    2026: "1c7462e42a99907ecdf58b05a90d5306",
}

CREATE_SQL = """
CREATE TABLE IF NOT EXISTS raw_eua_auction_eex (
  auction_date TEXT NOT NULL,
  auction_time TEXT,
  auction_name TEXT NOT NULL,
  contract TEXT,
  status TEXT,
  auction_price_eur REAL,
  min_bid_eur REAL, max_bid_eur REAL, mean_eur REAL, median_eur REAL,
  volume_tco2 INTEGER,
  total_bids_tco2 INTEGER,
  cover_ratio REAL,
  bidders INTEGER,
  successful_bidders INTEGER,
  revenue_eur REAL,
  zone TEXT,
  file_year INTEGER NOT NULL,
  loaded_at TEXT,
  PRIMARY KEY (auction_date, auction_name)
)
"""

# spec §3スキーマ列 <- 実測ヘッダ列名(列名ベースmapping。Country/Zoneのみ年で改名)
COL_MAP = {
    "auction_time": "Time",
    "contract": "Contract",
    "status": "Status",  # 2017-2019は列自体が無い -> None
    "auction_price_eur": "Auction Price €/tCO2",
    "min_bid_eur": "Minimum Bid €/tCO2",
    "max_bid_eur": "Maximum Bid €/tCO2",
    "mean_eur": "Mean €/tCO2",
    "median_eur": "Median €/tCO2",
    "volume_tco2": "Auction Volume tCO2",
    "total_bids_tco2": "Total Amount of Bids",
    "cover_ratio": "Cover Ratio",
    "bidders": "Total Number of Bidders",
    "successful_bidders": "Number of Successful Bidders",
    "revenue_eur": "Total Revenue €",
}
ZONE_ALIASES = ("Zone", "Country")


def _num(v, cast):
    if v is None or v == "":
        return None
    if isinstance(v, str) and v.strip().upper() in ("#N/A", "N/A", "-"):
        return None
    return cast(v)


def _read_xls(fp, year):
    wb = xlrd.open_workbook(str(fp))
    sh = wb.sheet_by_index(0)
    hdr_r = None
    for r in range(sh.nrows):
        vals = [str(sh.cell_value(r, c)).strip() for c in range(sh.ncols)]
        if "Date" in vals and "Auction Name" in vals:
            hdr_r = r
            break
    if hdr_r is None:
        raise RuntimeError(f"{fp.name}: header row not found (Date+Auction Name)")
    headers = [str(sh.cell_value(hdr_r, c)).strip() for c in range(sh.ncols)]
    idx = {h: i for i, h in enumerate(headers)}
    zone_col = next((z for z in ZONE_ALIASES if z in idx), None)

    rows = []
    for r in range(hdr_r + 1, sh.nrows):
        date_serial = sh.cell_value(r, idx["Date"])
        if date_serial == "" or date_serial is None:
            continue
        d = xlrd.xldate_as_datetime(date_serial, wb.datemode)
        time_serial = sh.cell_value(r, idx["Time"])
        t = xlrd.xldate_as_datetime(time_serial, wb.datemode).time().isoformat() if time_serial not in ("", None) else None
        row = _build_row(sh_get=lambda col: sh.cell_value(r, idx[col]) if col in idx else None,
                          date_str=d.strftime("%Y-%m-%d"), time_str=t,
                          zone_val=(sh.cell_value(r, idx[zone_col]) if zone_col else None),
                          status_present="Status" in idx, year=year)
        rows.append(row)
    return rows, sh.nrows, hdr_r


def _read_xlsx(fp, year):
    wb = openpyxl.load_workbook(str(fp), data_only=True)
    sh = wb.worksheets[0]
    hdr_r = None
    for r in range(1, sh.max_row + 1):
        vals = [str(sh.cell(row=r, column=c).value).strip() for c in range(1, sh.max_column + 1)]
        if "Date" in vals and "Auction Name" in vals:
            hdr_r = r
            break
    if hdr_r is None:
        raise RuntimeError(f"{fp.name}: header row not found (Date+Auction Name)")
    headers = [str(sh.cell(row=hdr_r, column=c).value).strip() for c in range(1, sh.max_column + 1)]
    idx = {h: i + 1 for i, h in enumerate(headers)}  # 1-index column number
    zone_col = next((z for z in ZONE_ALIASES if z in idx), None)

    rows = []
    for r in range(hdr_r + 1, sh.max_row + 1):
        date_val = sh.cell(row=r, column=idx["Date"]).value
        if date_val is None or date_val == "":
            continue
        d = date_val
        time_val = sh.cell(row=r, column=idx["Time"]).value if "Time" in idx else None
        t = time_val.time().isoformat() if hasattr(time_val, "time") else (str(time_val) if time_val not in (None, "") else None)
        row = _build_row(sh_get=lambda col: sh.cell(row=r, column=idx[col]).value if col in idx else None,
                          date_str=d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d),
                          time_str=t,
                          zone_val=(sh.cell(row=r, column=idx[zone_col]).value if zone_col else None),
                          status_present="Status" in idx, year=year)
        rows.append(row)
    return rows, sh.max_row, hdr_r


def _build_row(sh_get, date_str, time_str, zone_val, status_present, year):
    return (
        date_str,
        time_str,
        sh_get("Auction Name"),
        sh_get("Contract"),
        (str(sh_get("Status")) if status_present and sh_get("Status") not in (None, "") else None),
        _num(sh_get("Auction Price €/tCO2"), float),
        _num(sh_get("Minimum Bid €/tCO2"), float),
        _num(sh_get("Maximum Bid €/tCO2"), float),
        _num(sh_get("Mean €/tCO2"), float),
        _num(sh_get("Median €/tCO2"), float),
        _num(sh_get("Auction Volume tCO2"), int),
        _num(sh_get("Total Amount of Bids"), int),
        _num(sh_get("Cover Ratio"), float),
        _num(sh_get("Total Number of Bidders"), int),
        _num(sh_get("Number of Successful Bidders"), int),
        _num(sh_get("Total Revenue €"), float),
        (str(zone_val) if zone_val not in (None, "") else None),
        year,
    )


def verify_md5():
    for year, expected in EXPECTED_MD5.items():
        ext = "xls" if year <= 2019 else "xlsx"
        fp = SRC_DIR / f"eex_auction_{year}.{ext}"
        actual = hashlib.md5(fp.read_bytes()).hexdigest()
        if actual != expected:
            raise RuntimeError(f"md5 mismatch {fp.name}: expected {expected}, got {actual}. 素材改変疑い=投入中断。")
    print(f"[OK] md5 verified: {len(EXPECTED_MD5)}/10 files match SOURCES.md")


def load_all():
    all_rows = []
    per_file = {}
    for year in range(2017, 2027):
        ext = "xls" if year <= 2019 else "xlsx"
        fp = SRC_DIR / f"eex_auction_{year}.{ext}"
        rows, total_rows, hdr_r = (_read_xls(fp, year) if ext == "xls" else _read_xlsx(fp, year))
        per_file[year] = {"rows": len(rows), "total_rows": total_rows, "header_row": hdr_r}
        all_rows.extend(rows)
    return all_rows, per_file


def main():
    verify_md5()
    rows, per_file = load_all()
    print(f"total rows(全10ファイル合計)={len(rows)}")
    for year in sorted(per_file):
        pf = per_file[year]
        print(f"  {year}: data_rows={pf['rows']} total_rows={pf['total_rows']} header_row_0idx_or_1idx={pf['header_row']}")

    conn = sqlite3.connect(GODS)
    conn.execute(CREATE_SQL)

    existing = conn.execute("SELECT COUNT(*) FROM raw_eua_auction_eex").fetchone()[0]
    if existing:
        print(f"[SKIP-CLEAR] raw_eua_auction_eex already has {existing} rows. Clearing for idempotent reload.")
        conn.execute("DELETE FROM raw_eua_auction_eex")

    loaded_at = datetime.now().isoformat()
    cols = ("auction_date, auction_time, auction_name, contract, status, "
            "auction_price_eur, min_bid_eur, max_bid_eur, mean_eur, median_eur, "
            "volume_tco2, total_bids_tco2, cover_ratio, bidders, successful_bidders, "
            "revenue_eur, zone, file_year, loaded_at")
    placeholders = ", ".join(["?"] * 19)
    conn.executemany(
        f"INSERT INTO raw_eua_auction_eex ({cols}) VALUES ({placeholders})",
        [r + (loaded_at,) for r in rows],
    )
    conn.commit()

    total = conn.execute("SELECT COUNT(*), MIN(auction_date), MAX(auction_date) FROM raw_eua_auction_eex").fetchone()
    by_year = conn.execute(
        "SELECT file_year, COUNT(*) FROM raw_eua_auction_eex GROUP BY file_year ORDER BY file_year"
    ).fetchall()
    print(f"[DONE] raw_eua_auction_eex: rows={total[0]} range={total[1]}..{total[2]}")
    print(f"by file_year: {by_year}")
    conn.close()


if __name__ == "__main__":
    main()
